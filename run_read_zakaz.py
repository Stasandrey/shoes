#!/usr/bin/python3
# -*- coding: utf-8 -*-

#*******************************************************************************
#*                                                Разбор заказа обуви из файла XLSX                                        *
#*******************************************************************************

from PyQt5 import QtWidgets, QtGui#, QtCore
import read_zakaz

import sys
import openpyxl

class Read_Zakaz( QtWidgets.QWidget, read_zakaz.Ui_Read_Zakaz ):
    def __init__( self, database, parent = None   ):
        QtWidgets.QWidget.__init__( self, parent )
        self.setupUi( self )
        
        self.model = QtGui.QStandardItemModel( parent = self )   
        self.lstItems.setModel( self.model )
        
        
        
        self.btnOpen.clicked.connect( self.open )
        self.btnRead.clicked.connect( self.read )
        self.btnDel.clicked.connect( self.delete )
        self.btnClose.clicked.connect( self.close )
        
    def open( self ):
        dlg = QtWidgets.QFileDialog( filter = "Файлы заказов (*.xlsx)" )
        dlg.setFileMode( QtWidgets.QFileDialog.ExistingFiles )
        dlg.exec()
        name = dlg.selectedFiles()
        if len( name ) == 1:
          self.edtFilename.setText( name[0] )
    
    def read( self ):
        self.model.clear()
        item = QtGui.QStandardItem( "Разбор файла "+self.edtFilename.text() )
        self.model.appendRow( item )
        res = self.run( self.edtFilename.text() )
        self.model.clear()  
        for model in res:
            for size in model[1]:
                item = QtGui.QStandardItem( "Модель %s размер %s %s шт."%
                                           ( model[0], size[0], size[1] ) )
                self.model.appendRow( item )
        
        def delete( self ):
            if 

    def run( self, filename ):
        wb = openpyxl.load_workbook( filename )
        sh_names = wb.sheetnames
        if sh_names[0] != "Лист1":
            print( "Нет листа с именем 'Лист1'" )
            sys.exit()
        sheet = wb["Лист1"]
        if sheet["A1"].value.upper() != "МОДЕЛЬ":
            print( "Файл имеет неверный формат." )
            sys.exit()
        flag = True
        sizes = []
        #  Создаем список размеров
        for cellobj in sheet["B1":"Z1"]:
            for cell in cellobj:
                if flag == True:
                    if str( cell.value ).upper() != "ИТОГО":
                                        
                        sizes.append( [openpyxl.utils.get_column_letter( cell.column ), str( cell.value )] )
                    else:
                        flag = False
        max_size = len( sizes )
        models = []
        #  Создаем список моделей и номеров строк
        for cellobj in sheet["A1":"A1000"]:
            for cell in cellobj:
                m = str( cell.value ).upper()
                if  m != "NONE":
                    if m != "МОДЕЛЬ":
                        models.append( [m, cell.row] ) 
        res = []
        #  Создаем список вида [Модель, Размер, Количество]
        for model in models:
            res_m = []
            res_m.append( model[0] )
            res_s = []
            for i in range( max_size ):
                coord = "%s%s"%( sizes[i][0], model[1] )
                cell = sheet[coord].value
                if cell != None:
                    ms = [sizes[i][1], cell]
                    
                    res_s.append( ms )
            res_m.append( res_s )
            res.append( res_m )
        print( res )
        return res
        
        

if __name__ == "__main__":
    
    app = QtWidgets.QApplication( sys.argv )
    window = Read_Zakaz( 'DB' )
    window.show()
    sys.exit( app.exec_() )
